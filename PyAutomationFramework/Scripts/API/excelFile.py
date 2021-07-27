# excel read/write...
# pip install openpyxl     --- install this package

import openpyxl

book = openpyxl.load_workbook("C:\\Users\\shubhav\\PycharmProjects\\APITesting\\Utils\\pythonDemo.xlsx")
sheet = book.active  # Active sheet in the excel

# intialise empty dict

d = {}

cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = 'Hello'
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

# to print all the values
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)
    print("=======================")

# if yu want to print info based on say TC name say TC2
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TC2":
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)

#creating a dict from a search criteria .... say TC2 info yu need to put in a dict and print the same
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TC2":
        for j in range(1, sheet.max_column + 1):
            d[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(d)